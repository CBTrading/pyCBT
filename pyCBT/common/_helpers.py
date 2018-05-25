# TODO: define argument parser

def dump_data(*args, **kwargs):
    import os, sys
    from pyCBT.common.files import exist
    from openpyxl import load_workbook

    dataframe, = args
    if kwargs.get("save_to"):
        if "," in kwargs["save_to"]: filename, sheetname = kwargs["save_to"].split(",")
        else: filename, sheetname = kwargs.get("save_to"), "sheet_001"
        if filename.endswith(".xlsx") and exist(filename):
            # TODO: if the file exist, ask the user
            book = load_workbook(filename)
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                excel_writer.book = book
                if sheetname in book.sheetnames:
                    match = re.match("(\w+)_(\d+)", sheetname)
                    if not match: sheetname += "_{0:03d}"
                    else: sheetname = string.join([match.groups()[0], "{0:03d}"], "_")
                i = 1
                while sheetname.format(i) in book.sheetnames: i += 1
                sheetname = sheetname.format(i)
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
            book.close()
        elif filename.endswith(".xlsx"):
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
        else:
            # TODO: if the file exist, ask the user
            if filename.split(".")[-1] != "csv": filename += ".csv"
            dataframe.reset_index().to_csv(filename, index=False, line_terminator=os.linesep)
    else:
        dataframe.reset_index().to_csv(sys.stdout, index=False, line_terminator=os.linesep)

    return None
