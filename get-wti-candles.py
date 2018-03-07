#!/usr/bin/env python

def parse_args():
    import argparse, sys, pytz
    from datetime import datetime

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "resolution",
        help="Datetime resolution (example: M15, H1)",
        default="M15"
    )
    parser.add_argument(
        "from_date",
        metavar="from-datetime",
        help="Starting datetime in dataset"
    )
    parser.add_argument(
        "to_date",
        metavar="to-datetime",
        help="Closing datetime in dataset. Defaults to now in the given timezone"
    )
    parser.add_argument(
        "timezone",
        help="Timezone of 'from-datetime' & 'to-datetime' also alignment for the dataset"
    )
    parser.add_argument(
        "datetime_format",
        metavar="datetime-format",
        help="The datetime format of the dataset, example: RFC3339, UNIX, JSON, or any datetime format string.",
    )
    parser.add_argument(
        "--save-to", "-s",
        help="Stores the dataset in the given filename. Supported extensions are .csv (default if extension is missing) or .xlsx file. If .xlsx and a second argument (comma-separated) value is given, it is taken to be the name of the sheet."
    )
    # TODO: add interactive option
    # TODO: add verbose option

    # TODO: fix help string in save-to argument
    # TODO: fix help string in to-date
    args = parser.parse_args()

    if args.to_date.lower() == "now":
        args.to_date = datetime.now(tz=pytz.timezone(args.timezone)).isoformat()

    return args._get_args(), dict(args._get_kwargs())

def request_data(*args, **kwargs):
    from pyCBT.data.providers.oanda import account
    from pyCBT.data.providers.oanda import historical

    client = account.Client()
    candles = historical.Candles(
        client=client,
        instrument="WTICO_USD",
        resolution=kwargs.get("resolution"),
        from_date=kwargs.get("from_date"),
        to_date=kwargs.get("to_date"),
        datetime_fmt=kwargs.get("datetime_format"),
        timezone=kwargs.get("timezone")
    )
    return candles.as_dataframe()

def dump_data(*args, **kwargs):
    import os, sys, re, string
    import pandas as pd
    from pyCBT.tools.files import exist
    from openpyxl import load_workbook

    dataframe, = args
    if kwargs.get("save_to") is not None:
        if "," in kwargs["save_to"]: filename, sheetname = kwargs["save_to"].split(",")
        else: filename, sheetname = kwargs.get("save_to"), "sheet_001"
        if filename.endswith(".xlsx") and exist(filename):
            # TODO: if the file exist, ask the user
            book = load_workbook(filename)
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                excel_writer.book = book
                if sheetname in book.sheetnames:
                    match = re.match("(\w+)_(\d+)", sheetname)
                    if not match: sheetname += "_{0:03d}"
                    else: sheetname = string.join("_", [match.groups()[0], "{0:03d}"])
                i = 1
                while sheetname.format(i) in book.sheetnames: i += 1
                sheetname = sheetname.format(i)
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
            book.close()
        elif filename.endswith(".xlsx"):
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
        else:
            # TODO: if the file exist, ask the user
            df = dataframe.reset_index()
            df.to_csv(filename, index=False, line_terminator=os.linesep)

    ks = kwargs.keys()
    vs = kwargs.values()
    fmt_ln = [str(max(len(ks[i]),len(repr(vs[i])))) for i in xrange(len(ks))]
    fmt_st = string.join(["{%d:>%s}" % (i, fmt_ln[i]) for i in xrange(len(ks))], " ")

    print
    print "DATASET"
    print "-------"
    print >> sys.stdout, dataframe
    print
    print "COMAND LINE ARGUMENTS"
    print "---------------------"
    print fmt_st.format(*[ks[i] for i in xrange(len(ks))])
    print >> sys.stdout, fmt_st.format(*[vs[i] for i in xrange(len(vs))])
    print
    return None

if __name__ == "__main__":
    args, kwargs = parse_args()
    table = request_data(*args, **kwargs)
    dump_data(table, **kwargs)
